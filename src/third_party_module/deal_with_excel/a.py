# -*- coding:utf-8 -*-
from openpyxl import load_workbook

#   定义目录名
DIR_NAME = u"E:\\"

#   定义文件名
FILE_NAME = u"数据集成每日作业运行情况登记表_20161224.xlsx"


#   获取第一行的表头信息
def get_excel_title_list():
    templist = []
    for cx in range(ws.min_column, ws.max_column + 1):
        templist.append(ws.cell(row=1, column=cx).value)
    return templist


#   改变sheet页的名称
def rename_excel_sheet_name():
    wb.get_active_sheet().title = u"作业监控记录表"


#   保存excel文件
def save_excel():
    wb.save(DIR_NAME + FILE_NAME)


#   数据集成每日作业运行情况登记表_20161122.xlsx,处理11-14列
def deal_columns():
    for col_num in range(11, ws.max_column + 1):
        for row_num in range(1, 60):
            ws.cell(row=row_num, column=col_num).value = ""


wb = load_workbook(DIR_NAME + FILE_NAME)
sheetNames = wb.get_sheet_names()
ws = wb.get_sheet_by_name(sheetNames[0])
for x in get_excel_title_list():
    print
    x,
rename_excel_sheet_name()
deal_columns()
save_excel()